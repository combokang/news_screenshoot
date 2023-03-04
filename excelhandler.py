from openpyxl import load_workbook


print("""
版權聲明v.1.2

------------

* 本程式及其說明文件之智慧財產權，包括著作權屬於康博元(Po-Yuan Kang, hiromoto1995@gmail.com)個人所有

* 本程式僅授權精承永續企業顧問股份有限公司內部進行複製、使用，嚴禁授權範圍外之使用
 
------------
""")

while True:
    file_name = input("請右鍵貼上.xlsx檔案名稱並按下Enter鍵(不包含附檔名)：")
    try:
        wb = load_workbook(f"{file_name}.xlsx")
        break
    except FileNotFoundError as e:
        print("該檔案不存在，請重新輸入")

sheet_list = wb.sheetnames  # 讀取工作表清單

# for i in sheet_list:    # 遞迴讀取多個工作表
#     active_sheet = wb[sheet_list[i]]

print("\n開始讀取excel檔案")

url_dic = {}  # {新聞編號:超連結} 字典
first_sheet = wb[sheet_list[0]]  # 第一個工作表
for i in range(2, first_sheet.max_row+1):
    no = first_sheet[f"A{i}"].value  # 讀取編號
    if no != None:  # 如果有編號
        if first_sheet[f"H{i}"].hyperlink != None:  # 如果有超連結
            # 將編號和超連結加入字典
            url_dic[str(no)] = first_sheet[f"H{i}"].hyperlink.target
        else:
            print(f"編號: {no} 無超連結")
    else:   # 一沒有讀到編號即跳出迴圈
        break

print()
